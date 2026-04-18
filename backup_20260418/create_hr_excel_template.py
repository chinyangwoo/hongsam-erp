"""
홍삼스파 ERP - 인사기록카드 엑셀 템플릿 생성 스크립트
실행: python create_hr_excel_template.py
출력: 홍삼스파_인사기록카드_양식.xlsx
"""

import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

# ──────────────────────────────────────────────────────────────
# 색상 팔레트
# ──────────────────────────────────────────────────────────────
C_NAVY      = "0F172A"   # 헤더 배경 (ERP 메인 다크)
C_INDIGO    = "1E3A5F"   # 섹션 헤더
C_BLUE      = "1D4ED8"   # 강조 파란색
C_LIGHT_BG  = "EFF6FF"   # 입력 셀 배경
C_GRAY_BG   = "F8FAFC"   # 일반 배경
C_WHITE     = "FFFFFF"
C_GOLD      = "F59E0B"   # 필수항목 강조
C_RED_LIGHT = "FEF2F2"   # 공제항목 배경
C_GREEN_BG  = "F0FDF4"   # 지급항목 배경

# ──────────────────────────────────────────────────────────────
# 헬퍼 함수
# ──────────────────────────────────────────────────────────────
def make_border(style="thin"):
    s = Side(border_style=style, color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(border_style="medium", color="1D4ED8")
    n = Side(border_style="thin",   color="BBBBBB")
    return Border(left=t, right=t, top=t, bottom=t)

def header_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cell_write(ws, row, col, value, bold=False, font_size=10,
               fill_color=None, font_color="000000", h_align="center",
               v_align="center", wrap=False, border=True, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="맑은 고딕", bold=bold, size=font_size, color=font_color)
    c.alignment = align(h_align, v_align, wrap)
    if fill_color:
        c.fill = header_fill(fill_color)
    if border:
        c.border = make_border()
    if num_format:
        c.number_format = num_format
    return c

def label_cell(ws, row, col, text):
    """레이블 셀 (회색 배경)"""
    return cell_write(ws, row, col, text, bold=True, font_size=9,
                      fill_color="D1D5DB", font_color="1E293B", h_align="center")

def input_cell(ws, row, col, default_hint="", bg=C_LIGHT_BG, num_format=None):
    """입력 셀 (연한 파란 배경)"""
    return cell_write(ws, row, col, default_hint, bold=False, font_size=10,
                      fill_color=bg, font_color="374151", h_align="left",
                      v_align="center", wrap=True, num_format=num_format)

# ──────────────────────────────────────────────────────────────
# 워크북 생성
# ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1: 인사기록카드 (단일사원 상세)
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "인사기록카드"
ws1.sheet_view.showGridLines = False

# 열 너비 설정
col_widths = {
    1: 3,   # 여백
    2: 14,  3: 18,  4: 14,  5: 18,
    6: 14,  7: 18,  8: 14,  9: 18,
    10: 3   # 여백
}
for col, width in col_widths.items():
    ws1.column_dimensions[get_column_letter(col)].width = width

# 기본 행 높이
ws1.sheet_properties.defaultColWidth = 14
ws1.sheet_properties.defaultRowHeight = 22

# ─── 메인 타이틀 헤더 (병합 B1:I3) ───
ws1.merge_cells("B1:I3")
title_cell = ws1["B1"]
title_cell.value = "홍삼스파  인사기록카드"
title_cell.font = Font(name="맑은 고딕", bold=True, size=18, color=C_WHITE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
title_cell.fill = PatternFill(fill_type="solid", fgColor=C_NAVY)
ws1.row_dimensions[1].height = 14
ws1.row_dimensions[2].height = 22
ws1.row_dimensions[3].height = 14
thick_b = Side(border_style="medium", color="3B82F6")
for cell in [ws1["B1"], ws1["B2"], ws1["B3"]]:
    cell.border = Border(left=thick_b, top=thick_b, bottom=thick_b)
for cell in [ws1["I1"], ws1["I2"], ws1["I3"]]:
    cell.border = Border(right=thick_b, top=thick_b, bottom=thick_b)

# ─── 부제 ───
curr_row = 4
ws1.merge_cells(f"B{curr_row}:I{curr_row}")
sub = ws1[f"B{curr_row}"]
sub.value = f"작성일: {datetime.date.today().strftime('%Y년 %m월 %d일')}    ※ 굵은 음영칸에 입력하세요."
sub.font = Font(name="맑은 고딕", size=9, color="64748B")
sub.alignment = Alignment(horizontal="right", vertical="center")
sub.fill = header_fill("F1F5F9")
ws1.row_dimensions[curr_row].height = 16

# ═══ SECTION 1: 기본정보 ═══════════════════════════════════════
curr_row = 5
ws1.merge_cells(f"B{curr_row}:I{curr_row}")
sec = ws1[f"B{curr_row}"]
sec.value = "■  기본 인적사항"
sec.font = Font(name="맑은 고딕", bold=True, size=10, color=C_WHITE)
sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
sec.fill = header_fill(C_INDIGO)
ws1.row_dimensions[curr_row].height = 22

# Row 6-7: 사번 / 성명 / 사진안내
curr_row = 6
label_cell(ws1, curr_row, 2, "사  번 *")
input_cell(ws1, curr_row, 3, "예: 001")
label_cell(ws1, curr_row, 4, "성  명 *")
input_cell(ws1, curr_row, 5, "홍길동")
ws1.merge_cells(f"G{curr_row}:I{curr_row+5}")
photo = ws1.cell(row=curr_row, column=7)
photo.value = "📷 사진란\n(3×4cm)\n\n상단 [신규사원등록] 버튼으로\n사진 업로드"
photo.font = Font(name="맑은 고딕", size=9, color="6B7280")
photo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
photo.fill = header_fill("F9FAFB")
photo.border = make_border("medium")

curr_row = 7
label_cell(ws1, curr_row, 2, "생년월일 *")
input_cell(ws1, curr_row, 3, "YYYY-MM-DD")
label_cell(ws1, curr_row, 4, "성  별")
dv_gender = DataValidation(type="list", formula1='"남,여"', allow_blank=True)
ws1.add_data_validation(dv_gender)
inp = input_cell(ws1, curr_row, 5, "남")
dv_gender.add(inp)

curr_row = 8
label_cell(ws1, curr_row, 2, "주민번호 *")
input_cell(ws1, curr_row, 3, "앞6자리-*******")
label_cell(ws1, curr_row, 4, "국  적")
input_cell(ws1, curr_row, 5, "대한민국")

curr_row = 9
label_cell(ws1, curr_row, 2, "연 락 처 *")
input_cell(ws1, curr_row, 3, "010-0000-0000")
label_cell(ws1, curr_row, 4, "비상연락처")
input_cell(ws1, curr_row, 5, "010-0000-0000")

curr_row = 10
label_cell(ws1, curr_row, 2, "이메일")
ws1.merge_cells(f"C{curr_row}:E{curr_row}")
input_cell(ws1, curr_row, 3, "example@email.com")

curr_row = 11
label_cell(ws1, curr_row, 2, "주  소 *")
ws1.merge_cells(f"C{curr_row}:F{curr_row}")
input_cell(ws1, curr_row, 3, "도로명 주소 기재")
ws1.row_dimensions[curr_row].height = 24

# ═══ SECTION 2: 발령/재직정보 ════════════════════════════════
curr_row = 12
ws1.merge_cells(f"B{curr_row}:I{curr_row}")
sec2 = ws1[f"B{curr_row}"]
sec2.value = "■  발령 / 재직 정보"
sec2.font = Font(name="맑은 고딕", bold=True, size=10, color=C_WHITE)
sec2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
sec2.fill = header_fill(C_INDIGO)
ws1.row_dimensions[curr_row].height = 22

curr_row = 13
label_cell(ws1, curr_row, 2, "입 사 일 *")
input_cell(ws1, curr_row, 3, "YYYY-MM-DD")
label_cell(ws1, curr_row, 4, "부  서 *")
dv_dept = DataValidation(type="list",
    formula1='"지원팀,운영팀,식음료팀,공무팀,마케팅팀"', allow_blank=True)
ws1.add_data_validation(dv_dept)
i = input_cell(ws1, curr_row, 5, "운영팀")
dv_dept.add(i)
label_cell(ws1, curr_row, 6, "세  부  팀")
input_cell(ws1, curr_row, 7, "예: 2층 스파담당")

curr_row = 14
label_cell(ws1, curr_row, 2, "직 급 *")
dv_rank = DataValidation(type="list",
    formula1='"크루,주임,대리,과장,차장,부장,임원,대표이사"', allow_blank=True)
ws1.add_data_validation(dv_rank)
i = input_cell(ws1, curr_row, 3, "크루")
dv_rank.add(i)
label_cell(ws1, curr_row, 4, "고용형태 *")
dv_emp = DataValidation(type="list",
    formula1='"정규직,계약직,아르바이트,인턴"', allow_blank=True)
ws1.add_data_validation(dv_emp)
i = input_cell(ws1, curr_row, 5, "정규직")
dv_emp.add(i)
label_cell(ws1, curr_row, 6, "재직상태 *")
dv_status = DataValidation(type="list",
    formula1='"재직,휴직,퇴직"', allow_blank=True)
ws1.add_data_validation(dv_status)
i = input_cell(ws1, curr_row, 7, "재직")
dv_status.add(i)

curr_row = 15
label_cell(ws1, curr_row, 2, "계약만료일")
input_cell(ws1, curr_row, 3, "YYYY-MM-DD (계약직만)")
label_cell(ws1, curr_row, 4, "퇴 직 일")
input_cell(ws1, curr_row, 5, "해당없음")
label_cell(ws1, curr_row, 6, "퇴직사유")
input_cell(ws1, curr_row, 7, "")

# ═══ SECTION 3: 급여 정보 ════════════════════════════════════
curr_row = 16
ws1.merge_cells(f"B{curr_row}:I{curr_row}")
sec3 = ws1[f"B{curr_row}"]
sec3.value = "■  급여 / 계약 정보  (보안: 권한자만 열람)"
sec3.font = Font(name="맑은 고딕", bold=True, size=10, color=C_WHITE)
sec3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
sec3.fill = header_fill("1E3A5F")
ws1.row_dimensions[curr_row].height = 22

curr_row = 17
for c, (lbl, hint) in enumerate([
    ("계약연봉(원)","36000000"), ("기본급(원)","3000000"),
    ("식대(원)","200000"), ("연장/야간(원)","0")
], start=0):
    col = 2 + c * 2
    label_cell(ws1, curr_row, col, lbl)
    input_cell(ws1, curr_row, col+1, hint, bg=C_GREEN_BG, num_format="#,##0")

curr_row = 18
for c, (lbl, hint) in enumerate([
    ("국민연금(원)","135000"), ("건강보험(원)","106350"),
    ("고용보험(원)","27000"), ("갑근세/주민세","98000")
], start=0):
    col = 2 + c * 2
    label_cell(ws1, curr_row, col, lbl)
    input_cell(ws1, curr_row, col+1, hint, bg=C_RED_LIGHT, num_format="#,##0")

curr_row = 19
label_cell(ws1, curr_row, 2, "실수령액(원)")
ws1.merge_cells(f"C{curr_row}:E{curr_row}")
net = ws1.cell(row=curr_row, column=3)
net.value = "=C18+E18+G18+I18-C19-E19-G19-I19"  # placeholder formula comment
net.value = 0   # 수기입력용 (ERP에서 자동 계산)
net.font = Font(name="맑은 고딕", bold=True, size=11, color="1D4ED8")
net.alignment = Alignment(horizontal="right", vertical="center")
net.fill = header_fill("DBEAFE")
net.border = make_border("medium")
net.number_format = "#,##0"
ws1.merge_cells(f"F{curr_row}:I{curr_row}")
note = ws1.cell(row=curr_row, column=6)
note.value = "※ 실수령액은 ERP 업로드 시 자동계산됩니다."
note.font = Font(name="맑은 고딕", size=8, color="6B7280", italic=True)
note.alignment = Alignment(horizontal="left", vertical="center")
note.fill = header_fill("F1F5F9")

# ═══ SECTION 4: 학력 ══════════════════════════════════════
curr_row = 20
ws1.merge_cells(f"B{curr_row}:I{curr_row}")
sec4 = ws1[f"B{curr_row}"]
sec4.value = "■  학력사항"
sec4.font = Font(name="맑은 고딕", bold=True, size=10, color=C_WHITE)
sec4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
sec4.fill = header_fill(C_INDIGO)
ws1.row_dimensions[curr_row].height = 22

curr_row = 21
for col, (txt, w) in enumerate(zip(
    ["학교명","전공/계열","학위","졸업구분","졸업년월"],
    [4, 4, 2, 2, 3]
), start=2):
    label_cell(ws1, curr_row, col, txt)

curr_row = 22
for col, hint in enumerate(
    ["예: 홍삼대학교","경영학과","학사","졸업","2015-02"],
    start=2):
    if col <= 6:
        input_cell(ws1, curr_row, col, hint)

# ═══ SECTION 5: 경력사항 ═════════════════════════════════
curr_row = 23
ws1.merge_cells(f"B{curr_row}:I{curr_row}")
sec5 = ws1[f"B{curr_row}"]
sec5.value = "■  경력사항  (최대 3건 기재)"
sec5.font = Font(name="맑은 고딕", bold=True, size=10, color=C_WHITE)
sec5.alignment = Alignment(horizontal="left", vertical="center", indent=1)
sec5.fill = header_fill(C_INDIGO)
ws1.row_dimensions[curr_row].height = 22

curr_row = 24
for col, txt in enumerate(["기관/회사명","업종","직위/담당","재직기간(시작)","재직기간(종료)"], start=2):
    label_cell(ws1, curr_row, col, txt)
label_cell(ws1, curr_row, 7, "퇴직사유")

for r in range(25, 28):
    for col in [2,3,4,5,6,7]:
        input_cell(ws1, r, col, "")
    ws1.row_dimensions[r].height = 22

# ═══ SECTION 6: 가족사항 ═════════════════════════════════
curr_row = 28
ws1.merge_cells(f"B{curr_row}:I{curr_row}")
sec6 = ws1[f"B{curr_row}"]
sec6.value = "■  가족사항"
sec6.font = Font(name="맑은 고딕", bold=True, size=10, color=C_WHITE)
sec6.alignment = Alignment(horizontal="left", vertical="center", indent=1)
sec6.fill = header_fill(C_INDIGO)
ws1.row_dimensions[curr_row].height = 22

curr_row = 29
for col, txt in enumerate(["관계","성명","생년월일","직업","동거여부"], start=2):
    label_cell(ws1, curr_row, col, txt)
label_cell(ws1, curr_row, 7, "비고")

for r in range(30, 33):
    for col in [2,3,4,5,6,7]:
        input_cell(ws1, r, col, "")
    ws1.row_dimensions[r].height = 22

# ═══ SECTION 7: 시스템 권한 ═══════════════════════════════
curr_row = 33
ws1.merge_cells(f"B{curr_row}:I{curr_row}")
sec7 = ws1[f"B{curr_row}"]
sec7.value = "■  ERP 시스템 접근 권한  (관리자 작성)"
sec7.font = Font(name="맑은 고딕", bold=True, size=10, color=C_WHITE)
sec7.alignment = Alignment(horizontal="left", vertical="center", indent=1)
sec7.fill = header_fill("1E3A5F")
ws1.row_dimensions[curr_row].height = 22

curr_row = 34
for col, txt in enumerate(["메뉴 모듈","열람(읽기)","생성/수정(쓰기)"], start=2):
    label_cell(ws1, curr_row, col, txt)

menus = [
    "1. 통합 대시보드",
    "2. HR/근태 관리 (인사기록/급여)",
    "3. 영업 및 캘린더",
    "4. 재고 및 공무",
    "5. 트래픽 모니터링",
    "6. 문서 관리",
    "7. 전자결재",
]
dv_yn = DataValidation(type="list", formula1='"O,X"', allow_blank=True)
ws1.add_data_validation(dv_yn)

for i, menu in enumerate(menus):
    r = curr_row + 1 + i
    input_cell(ws1, r, 2, menu, bg="F1F5F9")
    for col in [3, 4]:
        c = input_cell(ws1, r, col, "O" if col == 3 else "X", bg=C_LIGHT_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        dv_yn.add(c)
    ws1.row_dimensions[r].height = 22

# ═══ SECTION 8: 특이사항 / 서명 ══════════════════════════
curr_row = 42
ws1.merge_cells(f"B{curr_row}:I{curr_row}")
sec8 = ws1[f"B{curr_row}"]
sec8.value = "■  특이사항 / 관리자 메모"
sec8.font = Font(name="맑은 고딕", bold=True, size=10, color=C_WHITE)
sec8.alignment = Alignment(horizontal="left", vertical="center", indent=1)
sec8.fill = header_fill(C_INDIGO)
ws1.row_dimensions[curr_row].height = 22

curr_row = 43
ws1.merge_cells(f"B{curr_row}:I{curr_row+1}")
memo = ws1[f"B{curr_row}"]
memo.value = ""
memo.fill = header_fill(C_LIGHT_BG)
memo.border = make_border()
memo.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
ws1.row_dimensions[curr_row].height = 30
ws1.row_dimensions[curr_row+1].height = 30

curr_row = 45
ws1.merge_cells(f"F{curr_row}:I{curr_row+1}")
sign = ws1[f"F{curr_row}"]
sign.value = "등록 확인자 서명:                           (인)"
sign.font = Font(name="맑은 고딕", size=10, color="374151")
sign.alignment = Alignment(horizontal="center", vertical="center")
sign.border = make_border("medium")
ws1.row_dimensions[curr_row].height = 28

# 인쇄 설정
ws1.page_setup.orientation = "portrait"
ws1.page_setup.paperSize = 9  # A4
ws1.page_margins.left = 0.5
ws1.page_margins.right = 0.5
ws1.page_margins.top = 0.75
ws1.page_margins.bottom = 0.75
ws1.print_title_rows = "1:4"

ws1.freeze_panes = "B6"

# ══════════════════════════════════════════════════════════════
# SHEET 2: 일괄등록 (다수 사원 한번에 업로드)
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("일괄등록 (ERP 업로드용)")
ws2.sheet_view.showGridLines = False

# 헤더
headers = [
    ("사번*",           "emp_id",           8),
    ("성명*",           "name",             10),
    ("생년월일*",       "birth_date",       12),
    ("성별",            "gender",           6),
    ("주민번호앞6자리", "id_front",         14),
    ("연락처*",         "phone",            14),
    ("비상연락처",      "emergency_phone",  14),
    ("이메일",          "email",            20),
    ("주소*",           "address",          30),
    ("입사일*",         "hire_date",        12),
    ("부서*",           "department",       10),
    ("세부팀",          "team_detail",      14),
    ("직급*",           "rank",             8),
    ("고용형태*",       "emp_type",         10),
    ("재직상태*",       "status",           8),
    ("계약만료일",      "contract_end",     12),
    ("계약연봉(원)",    "annual_salary",    14),
    ("기본급(원)*",     "base_salary",      12),
    ("식대(원)",        "meal_allowance",   10),
    ("연장야간(원)",    "ot_allowance",     10),
    ("국민연금(원)",    "national_pension", 12),
    ("건강보험(원)",    "health_insurance", 12),
    ("고용보험(원)",    "employment_ins",   12),
    ("갑근세주민세(원)","income_tax",       14),
    ("가족사항",        "family_info",      16),
    ("특이사항",        "notes",            20),
    ("열람권한",        "perm_read",        10),
    ("수정권한",        "perm_write",       10),
]

# 타이틀
ws2.merge_cells("A1:AB1")
t = ws2["A1"]
t.value = "홍삼스파 ERP  ─  인사기록카드 일괄등록 양식  (이 시트만 업로드하세요)"
t.font = Font(name="맑은 고딕", bold=True, size=13, color=C_WHITE)
t.alignment = Alignment(horizontal="center", vertical="center")
t.fill = header_fill(C_NAVY)
ws2.row_dimensions[1].height = 32

ws2.merge_cells("A2:AB2")
guide = ws2["A2"]
guide.value = (
    "⚠ 필수(*) 항목은 반드시 입력해야 ERP 등록이 가능합니다.  "
    "| 날짜 형식: YYYY-MM-DD  "
    "| 부서: 지원팀/운영팀/식음료팀/공무팀/마케팅팀  "
    "| 직급: 크루/주임/대리/과장/차장/부장/임원/대표이사  "
    "| 고용형태: 정규직/계약직/아르바이트/인턴  "
    "| 재직상태: 재직/휴직/퇴직  "
    "| 권한: O / X"
)
guide.font = Font(name="맑은 고딕", size=8.5, color="374151")
guide.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
guide.fill = header_fill("FEF9C3")
ws2.row_dimensions[2].height = 36

# 컬럼 헤더 (한글) - Row 3
# 컬럼 헤더 (영문 필드명) - Row 4
ws2.row_dimensions[3].height = 24
ws2.row_dimensions[4].height = 18

for col_idx, (kor_name, eng_name, width) in enumerate(headers, start=1):
    col_letter = get_column_letter(col_idx)
    ws2.column_dimensions[col_letter].width = width

    # 한글 헤더
    c3 = ws2.cell(row=3, column=col_idx, value=kor_name)
    is_required = "*" in kor_name
    c3.font = Font(name="맑은 고딕", bold=True, size=9.5,
                   color=C_WHITE if is_required else "E2E8F0")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.fill = header_fill(C_BLUE if is_required else "2563EB")
    c3.border = make_border()

    # 영문 필드명 (ERP가 읽는 키)
    c4 = ws2.cell(row=4, column=col_idx, value=eng_name)
    c4.font = Font(name="Consolas", size=7.5, color="6B7280", italic=True)
    c4.alignment = Alignment(horizontal="center", vertical="center")
    c4.fill = header_fill("F8FAFC")
    c4.border = make_border()

# 샘플 데이터 2행
samples = [
    # 사번, 성명, 생년월일, 성별, 주민번호앞6, 연락처, 비상, 이메일,
    # 주소, 입사일, 부서, 세부팀, 직급, 고용형태, 재직상태, 계약만료,
    # 계약연봉, 기본급, 식대, 연장야간, 국민연금, 건강보험, 고용보험, 갑근세,
    # 가족사항, 특이사항, 열람권한, 수정권한
    [
        "105","홍길동","1990-05-21","남","900521","010-1234-5678","010-9876-5432","hong@email.com",
        "서울시 강남구 테헤란로 123","2025-01-15","운영팀","2층 스파담당","크루","정규직","재직","",
        38000000, 3000000, 200000, 450000, 135000, 106350, 27000, 98000,
        "기혼 / 자녀 1명","우수직원 표창 2026.03","O","X"
    ],
    [
        "021","김지원","1988-11-03","여","881103","010-2345-6789","010-8765-4321","kim@email.com",
        "서울시 서초구 서초대로 456","2022-03-02","지원팀","인사/총무","대리","정규직","재직","",
        48000000, 3800000, 200000, 0, 171000, 134580, 34200, 185000,
        "미혼","팀장 큐레이터 역할 수행","O","O"
    ],
]

ws2.row_dimensions[5].height = 22
ws2.row_dimensions[6].height = 22

fill_even = header_fill("EFF6FF")
fill_odd  = header_fill(C_WHITE)

for r_idx, row_data in enumerate(samples, start=5):
    for c_idx, val in enumerate(row_data, start=1):
        c = ws2.cell(row=r_idx, column=c_idx, value=val)
        c.font = Font(name="맑은 고딕", size=9.5, color="1E293B")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill_even if r_idx % 2 == 1 else fill_odd
        c.border = make_border()
        if c_idx in [17, 18, 19, 20, 21, 22, 23, 24]:  # 금액 컬럼
            c.number_format = "#,##0"

# 입력 안내 행 (7번째~)
ws2.merge_cells("A7:AB7")
placeholder = ws2["A7"]
placeholder.value = "← 위 예시를 참고하여 아래 8행부터 데이터를 입력하세요. 7행 이 행은 삭제해도 됩니다."
placeholder.font = Font(name="맑은 고딕", size=9, color="9CA3AF", italic=True)
placeholder.alignment = Alignment(horizontal="left", vertical="center", indent=2)
placeholder.fill = header_fill("FFFBEB")
ws2.row_dimensions[7].height = 20

# 빈 입력행 10개
for r in range(8, 18):
    ws2.row_dimensions[r].height = 22
    for c in range(1, len(headers)+1):
        cell = ws2.cell(row=r, column=c, value="")
        cell.fill = fill_even if r % 2 == 0 else fill_odd
        cell.border = make_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

ws2.freeze_panes = "A5"
ws2.page_setup.orientation = "landscape"
ws2.page_setup.paperSize = 9
ws2.page_setup.fitToPage = True
ws2.page_setup.fitToWidth = 1

# ══════════════════════════════════════════════════════════════
# SHEET 3: 코드표 (참조용)
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("코드표 (참조용)")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 2
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 30
ws3.column_dimensions["D"].width = 4
ws3.column_dimensions["E"].width = 18
ws3.column_dimensions["F"].width = 30

ws3.merge_cells("B1:F1")
code_title = ws3["B1"]
code_title.value = "홍삼스파 ERP  ─  코드 참조표  (수정 금지)"
code_title.font = Font(name="맑은 고딕", bold=True, size=12, color=C_WHITE)
code_title.alignment = Alignment(horizontal="center", vertical="center")
code_title.fill = header_fill(C_NAVY)
ws3.row_dimensions[1].height = 28

code_tables = [
    (3, "부서코드", [
        ("지원팀",    "경영지원, 인사, 총무"),
        ("운영팀",    "스파 1~3층 운영"),
        ("식음료팀",  "루프탑/카페 F&B 운영"),
        ("공무팀",    "시설/전기/기계"),
        ("마케팅팀",  "SNS·예약 마케팅"),
    ]),
    (10, "직급코드", [
        ("크루",      "일반 팀원"),
        ("주임",      "주임급"),
        ("대리",      "대리급"),
        ("과장",      "과장급"),
        ("차장",      "차장급"),
        ("부장",      "부장급"),
        ("임원",      "이사·상무 등"),
        ("대표이사",  "CEO"),
    ]),
    (20, "고용형태", [
        ("정규직",    "무기계약, 4대보험 전부"),
        ("계약직",    "기간제 근로자"),
        ("아르바이트","단기/시간제 근로"),
        ("인턴",      "수습기간 포함"),
    ]),
    (26, "재직상태", [
        ("재직",  "정상 근무 중"),
        ("휴직",  "육아·병가·기타 휴직"),
        ("퇴직",  "퇴사 처리 완료"),
    ]),
]

for start_row, table_name, rows in code_tables:
    ws3.merge_cells(f"B{start_row}:C{start_row}")
    hdr = ws3[f"B{start_row}"]
    hdr.value = table_name
    hdr.font = Font(name="맑은 고딕", bold=True, size=10, color=C_WHITE)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    hdr.fill = header_fill(C_INDIGO)
    ws3.row_dimensions[start_row].height = 22

    for i, (code, desc) in enumerate(rows, start=1):
        r = start_row + i
        c1 = ws3.cell(row=r, column=2, value=code)
        c1.font = Font(name="맑은 고딕", bold=True, size=9.5)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.fill = header_fill(C_LIGHT_BG)
        c1.border = make_border()

        c2 = ws3.cell(row=r, column=3, value=desc)
        c2.font = Font(name="맑은 고딕", size=9.5, color="374151")
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c2.fill = header_fill(C_WHITE)
        c2.border = make_border()
        ws3.row_dimensions[r].height = 20

# ──────────────────────────────────────────────────────────────
# 저장
# ──────────────────────────────────────────────────────────────
output_path = "홍삼스파_인사기록카드_양식.xlsx"
wb.save(output_path)
import sys
out = sys.stdout
out.write("[완료] 엑셀 템플릿 생성 완료: " + output_path + "\n")
out.write("   - Sheet 1: 인사기록카드 (개인 상세 양식, 인쇄용)\n")
out.write("   - Sheet 2: 일괄등록 (ERP 업로드용)\n")
out.write("   - Sheet 3: 코드표  (부서/직급/고용형태 참조)\n")
